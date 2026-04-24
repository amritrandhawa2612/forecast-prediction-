#!/usr/bin/env python3
"""


Usage:
    python sales_forecast.py                         
    python sales_forecast.py path/to/your_file.xlsx  


"""

import sys
import json
import math
import tempfile
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)


# ── Configuration 

DEFAULT_FILE = "2506_6536_BWA.xlsx"   # change to your file name / path
STORE_ID     = 6536
STATE_CODE   = "HH"                    # Hamburg — change for other stores
GROWTH_RATE  = None                    # None = auto-detect from Ist vs VJ data


# ── Data extraction 

def load_workbook(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def get_vorjahr(wb):
    """Load daily prior-year actuals from TagesUmsatzVorjahr sheet."""
    ws = wb["TagesUmsatzVorjahr"]
    data = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) > 4 and isinstance(row[2], datetime) and isinstance(row[4], (int, float)):
            data[row[2].strftime("%Y-%m-%d")] = round(float(row[4]), 2)
    return data


def get_holidays(wb, state_col_name=STATE_CODE):
    """Load public holidays for the given state from Feiertage sheet."""
    ws = wb["Feiertage"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 4 (index 3) contains state codes
    header = rows[3] if len(rows) > 3 else []
    state_col = next((i for i, v in enumerate(header) if str(v) == state_col_name), None)
    if state_col is None:
        print(f"WARNING: State column '{state_col_name}' not found in Feiertage sheet.")
        return {}
    holidays = {}
    for row in rows[4:]:
        if isinstance(row[1], datetime) and state_col < len(row) and row[state_col] == 1:
            holidays[row[1].strftime("%Y-%m-%d")] = str(row[3]) if row[3] else "Feiertag"
    return holidays


def get_schulferien(wb, state_col_name=STATE_CODE):
    """Load school break dates for the given state from Schulferien sheet."""
    ws = wb["Schulferien"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 3 (index 2) contains state codes
    header = rows[2] if len(rows) > 2 else []
    state_col = next((i for i, v in enumerate(header) if str(v) == state_col_name), None)
    if state_col is None:
        print(f"WARNING: State column '{state_col_name}' not found in Schulferien sheet.")
        return {}
    schulferien = {}
    for row in rows[3:]:
        if isinstance(row[1], datetime) and state_col < len(row) and row[state_col]:
            schulferien[row[1].strftime("%Y-%m-%d")] = str(row[state_col])
    return schulferien


def get_ist_values(wb):
    """Load existing Ist Netto and original Forecast from Umsatz Forecast sheet."""
    ws = wb["Umsatz Forecast"]
    ist = {}
    for row in ws.iter_rows(values_only=True):
        # col index 3=Datum, 7=VJ Netto, 8=Forecast netto, 10=Ist Netto
        if len(row) > 10 and isinstance(row[3], datetime):
            ds = row[3].strftime("%Y-%m-%d")
            ist[ds] = {
                "vj_fc":     round(float(row[7]),  2) if isinstance(row[7],  (int, float)) else None,
                "fc_orig":   round(float(row[8]),  2) if isinstance(row[8],  (int, float)) else None,
                "ist":       round(float(row[10]), 2) if isinstance(row[10], (int, float)) else None,
                "feiertag":  str(row[0]) if row[0] else None,
                "schulferien": str(row[1]) if row[1] else None,
            }
    return ist


# ── Forecast model 

def build_model(vorjahr, holidays, schulferien):
    """
    Learn day-of-week multipliers, holiday effect, and school-break effect
    from the prior year actuals.
    """
    day_groups = defaultdict(list)
    for ds, val in vorjahr.items():
        d = datetime.strptime(ds, "%Y-%m-%d")
        if ds not in holidays and ds not in schulferien:
            day_groups[d.weekday()].append(val)

    day_avgs = {dow: (sum(v)/len(v) if v else 0) for dow, v in day_groups.items()}
    overall_avg = sum(day_avgs.values()) / max(len(day_avgs), 1)
    day_mult = {dow: (avg / overall_avg if overall_avg else 1.0)
                for dow, avg in day_avgs.items()}

    def effect(group_dates):
        factors = []
        for ds in group_dates:
            if ds in vorjahr:
                d = datetime.strptime(ds, "%Y-%m-%d")
                base = day_avgs.get(d.weekday(), overall_avg)
                if base > 0:
                    factors.append(vorjahr[ds] / base)
        return sum(factors)/len(factors) if factors else 1.0

    holiday_mult  = effect(holidays)
    sf_mult       = effect([ds for ds in schulferien if ds not in holidays])

    return day_mult, day_avgs, overall_avg, holiday_mult, sf_mult


def compute_growth_rate(ist_data, vorjahr):
    """Derive actual growth rate from Ist vs VJ where both exist."""
    factors = []
    for ds, rec in ist_data.items():
        if rec["ist"] and ds in vorjahr and vorjahr[ds] > 0:
            factors.append(rec["ist"] / vorjahr[ds])
    return round(sum(factors)/len(factors), 4) if factors else 1.04


def generate_forecast(year, vorjahr, holidays, schulferien, ist_data,
                      day_mult, day_avgs, overall_avg,
                      holiday_mult, sf_mult, growth_rate):
    """Build a daily forecast for the given year."""
    DAY_NAMES = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
    results = []
    cur = datetime(year, 1, 1)
    end = datetime(year, 12, 31)

    while cur <= end:
        ds = cur.strftime("%Y-%m-%d")
        # Map to prior year by same weekday (52-week shift keeps weekdays aligned)
        vj_date = cur - timedelta(days=364)
        vj_ds   = vj_date.strftime("%Y-%m-%d")
        vj_val  = vorjahr.get(vj_ds) or vorjahr.get((cur - timedelta(days=365)).strftime("%Y-%m-%d"))

        dow = cur.weekday()
        base = (vj_val * growth_rate) if vj_val else (day_avgs.get(dow, overall_avg) * growth_rate)

        if ds in holidays:
            mult = holiday_mult
        elif ds in schulferien:
            mult = sf_mult
        else:
            mult = 1.0

        ist_rec = ist_data.get(ds, {})
        results.append({
            "date":       ds,
            "dow":        dow,
            "dowName":    DAY_NAMES[dow],
            "vj":         round(vj_val, 2) if vj_val else None,
            "forecast":   round(base * mult),
            "ist":        ist_rec.get("ist"),
            "fc_orig":    ist_rec.get("fc_orig"),
            "feiertag":   holidays.get(ds),
            "schulferien": schulferien.get(ds),
        })
        cur += timedelta(days=1)
    return results


# ── Monthly aggregates 

def monthly_summary(results):
    MONTH_NAMES = ["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez"]
    monthly = defaultdict(lambda: {"forecast":0,"vj":0,"ist":0,"ist_days":0,"days":0})
    for r in results:
        m = r["date"][:7]
        monthly[m]["forecast"] += r["forecast"]
        monthly[m]["days"]     += 1
        if r["vj"]:  monthly[m]["vj"] += r["vj"]
        if r["ist"]: monthly[m]["ist"] += r["ist"]; monthly[m]["ist_days"] += 1
    out = []
    for i, (mk, mv) in enumerate(sorted(monthly.items())):
        out.append({
            "month":    MONTH_NAMES[i],
            "key":      mk,
            "forecast": round(mv["forecast"]),
            "vj":       round(mv["vj"]),
            "ist":      round(mv["ist"]) if mv["ist_days"] > 0 else None,
            "ist_days": mv["ist_days"],
            "days":     mv["days"],
        })
    return out


# ── HTML dashboard 

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Sales Forecast — HAMBURG Jungfernstieg 6536</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Syne:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
  :root{--bg:#0d0f14;--surface:#13161e;--surface2:#1a1e28;--border:#252a38;--accent:#e8f04a;--accent2:#4af0a8;--accent3:#f04a7a;--text:#e8eaf0;--muted:#6b7280;--gold:#f5c842;}
  *{box-sizing:border-box;margin:0;padding:0;}
  body{background:var(--bg);color:var(--text);font-family:'DM Mono',monospace;font-size:13px;min-height:100vh;}
  .header{display:flex;align-items:center;justify-content:space-between;padding:20px 28px;border-bottom:1px solid var(--border);background:var(--surface);}
  .header-left{display:flex;align-items:center;gap:16px;}
  .logo{width:36px;height:36px;background:var(--accent);border-radius:6px;display:flex;align-items:center;justify-content:center;font-family:'Syne',sans-serif;font-weight:800;font-size:14px;color:#0d0f14;}
  .header-title{font-family:'Syne',sans-serif;font-weight:700;font-size:15px;color:var(--text);}
  .header-sub{font-size:11px;color:var(--muted);margin-top:2px;letter-spacing:.05em;}
  .badge{background:var(--surface2);border:1px solid var(--border);border-radius:20px;padding:4px 12px;font-size:11px;color:var(--muted);}
  .main{padding:24px 28px;}
  .kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px;}
  .kpi{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:16px 18px;position:relative;overflow:hidden;}
  .kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;}
  .kpi.yellow::before{background:var(--accent);}
  .kpi.green::before{background:var(--accent2);}
  .kpi.red::before{background:var(--accent3);}
  .kpi.gold::before{background:var(--gold);}
  .kpi-label{font-size:10px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);margin-bottom:8px;}
  .kpi-val{font-family:'Syne',sans-serif;font-size:22px;font-weight:700;color:var(--text);}
  .kpi-sub{font-size:11px;color:var(--muted);margin-top:4px;}
  .up{color:var(--accent2);} .down{color:var(--accent3);}
  .cols{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px;}
  .card{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;}
  .card-title{font-family:'Syne',sans-serif;font-weight:600;font-size:12px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:16px;}
  .chart-wrap{position:relative;height:180px;}
  .month-pills{display:flex;gap:4px;flex-wrap:wrap;margin-bottom:16px;}
  .mpill{padding:4px 10px;border-radius:4px;font-size:11px;cursor:pointer;border:1px solid var(--border);background:transparent;color:var(--muted);transition:all .15s;letter-spacing:.03em;}
  .mpill:hover{border-color:var(--accent);color:var(--accent);}
  .mpill.active{background:var(--accent);color:#0d0f14;border-color:var(--accent);font-weight:500;}
  .table-wrap{overflow-x:auto;max-height:420px;overflow-y:auto;}
  .table-wrap::-webkit-scrollbar{width:4px;height:4px;}
  .table-wrap::-webkit-scrollbar-track{background:var(--surface);}
  .table-wrap::-webkit-scrollbar-thumb{background:var(--border);border-radius:2px;}
  table{width:100%;border-collapse:collapse;font-size:12px;}
  th{font-size:10px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);padding:8px 10px;text-align:left;border-bottom:1px solid var(--border);position:sticky;top:0;background:var(--surface);z-index:1;}
  td{padding:7px 10px;border-bottom:1px solid rgba(37,42,56,.5);}
  tr:hover td{background:var(--surface2);}
  .num{text-align:right;font-variant-numeric:tabular-nums;}
  .tag-f{background:rgba(240,74,122,.15);color:#f04a7a;border:1px solid rgba(240,74,122,.3);border-radius:3px;padding:1px 6px;font-size:10px;white-space:nowrap;}
  .tag-s{background:rgba(74,240,168,.1);color:#4af0a8;border:1px solid rgba(74,240,168,.25);border-radius:3px;padding:1px 6px;font-size:10px;white-space:nowrap;}
  .tag-w{background:rgba(232,240,74,.1);color:#e8f04a;border:1px solid rgba(232,240,74,.2);border-radius:3px;padding:1px 6px;font-size:10px;}
  .tag-d{background:rgba(107,114,128,.15);color:var(--muted);border:1px solid rgba(107,114,128,.2);border-radius:3px;padding:1px 6px;font-size:10px;}
  .dot-ist{display:inline-block;width:6px;height:6px;border-radius:50%;background:var(--accent2);margin-right:4px;vertical-align:middle;}
  .full-chart{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;margin-bottom:20px;}
  .full-chart-wrap{position:relative;height:200px;}
  .filter-row{display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap;}
  .filter-btn{padding:5px 12px;border-radius:4px;font-size:11px;cursor:pointer;border:1px solid var(--border);background:transparent;color:var(--muted);font-family:'DM Mono',monospace;transition:all .15s;}
  .filter-btn.active{background:var(--surface2);border-color:var(--accent);color:var(--accent);}
  .legend{display:flex;gap:16px;margin-bottom:12px;}
  .leg-item{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--muted);}
  .leg-dot{width:10px;height:3px;border-radius:2px;}
  .stats-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;}
  .stat-item{background:var(--surface2);border-radius:6px;padding:10px 12px;}
  .stat-label{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;}
  .stat-val{font-family:'Syne',sans-serif;font-size:14px;font-weight:600;color:var(--text);}
  .export-btn{background:var(--surface2);border:1px solid var(--border);color:var(--text);padding:8px 16px;border-radius:6px;font-family:'DM Mono',monospace;font-size:12px;cursor:pointer;transition:all .15s;}
  .export-btn:hover{border-color:var(--accent);color:var(--accent);}
  @media(max-width:900px){.kpi-row{grid-template-columns:repeat(2,1fr);}.cols{grid-template-columns:1fr;}.stats-grid{grid-template-columns:repeat(2,1fr);}}
</style>
</head>
<body>
<div class="header">
  <div class="header-left">
    <div class="logo">SF</div>
    <div>
      <div class="header-title">HAMBURG Jungfernstieg</div>
      <div class="header-sub">STORE __STORE_ID__ · SALES FORECAST __YEAR__</div>
    </div>
  </div>
  <div style="display:flex;gap:8px;align-items:center;">
    <span class="badge">Modell: VJ × Wachstum × Wochentag</span>
    <button class="export-btn" onclick="exportCSV()">↓ CSV Export</button>
  </div>
</div>
<div class="main">
  <div class="kpi-row">
    <div class="kpi yellow"><div class="kpi-label">Jahresprognose __YEAR__</div><div class="kpi-val" id="kpi-total">—</div><div class="kpi-sub">Netto Umsatz gesamt</div></div>
    <div class="kpi green"><div class="kpi-label">Ø Tagesumsatz FC</div><div class="kpi-val" id="kpi-avg">—</div><div class="kpi-sub">Alle 365 Tage</div></div>
    <div class="kpi red"><div class="kpi-label">vs. Vorjahr</div><div class="kpi-val" id="kpi-vs">—</div><div class="kpi-sub" id="kpi-vs-sub">—</div></div>
    <div class="kpi gold"><div class="kpi-label">Ist (verfügbar)</div><div class="kpi-val" id="kpi-ist">—</div><div class="kpi-sub" id="kpi-ist-sub">—</div></div>
  </div>
  <div class="full-chart">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;">
      <div class="card-title" style="margin:0;">Jahresverlauf — Vorjahr vs. Forecast vs. Ist</div>
      <div class="legend">
        <div class="leg-item"><div class="leg-dot" style="background:#6b7280;"></div> Vorjahr</div>
        <div class="leg-item"><div class="leg-dot" style="background:#e8f04a;"></div> Forecast</div>
        <div class="leg-item"><div class="leg-dot" style="background:#4af0a8;"></div> Ist</div>
      </div>
    </div>
    <div class="full-chart-wrap"><canvas id="yearChart"></canvas></div>
  </div>
  <div class="cols">
    <div class="card">
      <div class="card-title">Monatliche Übersicht</div>
      <div class="chart-wrap"><canvas id="monthChart"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title">Modell-Parameter</div>
      <div class="stats-grid">
        <div class="stat-item"><div class="stat-label">Wachstum VJ</div><div class="stat-val" id="s-growth">—</div></div>
        <div class="stat-item"><div class="stat-label">Feiertag-Faktor</div><div class="stat-val" id="s-holiday">—</div></div>
        <div class="stat-item"><div class="stat-label">Schulferien-Faktor</div><div class="stat-val" id="s-sf">—</div></div>
        <div class="stat-item"><div class="stat-label">Bester Tag</div><div class="stat-val">Sa ×<span id="s-sat">—</span></div></div>
        <div class="stat-item"><div class="stat-label">Schwächster Tag</div><div class="stat-val">Mo ×<span id="s-mon">—</span></div></div>
        <div class="stat-item"><div class="stat-label">Feiertage HH</div><div class="stat-val" id="s-fdays">—</div></div>
      </div>
      <div style="margin-top:14px;">
        <div class="stat-label" style="margin-bottom:8px;">Wochentag-Multiplikatoren</div>
        <div id="day-bars" style="display:flex;gap:4px;align-items:flex-end;height:50px;"></div>
        <div style="display:flex;gap:4px;margin-top:4px;" id="day-labels"></div>
      </div>
    </div>
  </div>
  <div class="card">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;">
      <div class="card-title" style="margin:0;">Tagesprognose</div>
      <div class="month-pills" id="month-pills" style="margin:0;"></div>
    </div>
    <div class="filter-row">
      <span style="font-size:11px;color:var(--muted);">Filter:</span>
      <button class="filter-btn active" onclick="setFilter('all',this)">Alle</button>
      <button class="filter-btn" onclick="setFilter('feiertag',this)">Feiertage</button>
      <button class="filter-btn" onclick="setFilter('schulferien',this)">Schulferien</button>
      <button class="filter-btn" onclick="setFilter('weekend',this)">Wochenende</button>
      <button class="filter-btn" onclick="setFilter('ist',this)">Mit Ist</button>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>Datum</th><th>Tag</th><th>Typ</th>
          <th class="num">Vorjahr €</th><th class="num">Forecast €</th>
          <th class="num">Ist €</th><th class="num">FC vs VJ</th><th class="num">Ist vs FC</th>
        </tr></thead>
        <tbody id="day-tbody"></tbody>
      </table>
    </div>
  </div>
</div>
<script>
const RAW=__RESULTS__;
const MONTHLY=__MONTHLY__;
const META=__META__;
const fmt=v=>v!=null?Math.round(v).toLocaleString('de-DE'):'—';
const fmtPct=v=>(v>=0?'+':'')+(v*100).toFixed(1)+'%';
const DAYS=['Mo','Di','Mi','Do','Fr','Sa','So'];
let activeMonth='all',activeFilter='all';
const totalFC=RAW.reduce((s,r)=>s+r.forecast,0);
const totalVJ=RAW.reduce((s,r)=>s+(r.vj||0),0);
const totalIst=RAW.filter(r=>r.ist).reduce((s,r)=>s+r.ist,0);
const istRows=RAW.filter(r=>r.ist);
document.getElementById('kpi-total').textContent='€'+fmt(totalFC);
document.getElementById('kpi-avg').textContent='€'+fmt(totalFC/365);
const vsVJ=(totalFC-totalVJ)/totalVJ;
document.getElementById('kpi-vs').textContent=fmtPct(vsVJ);
document.getElementById('kpi-vs').className='kpi-val '+(vsVJ>=0?'up':'down');
document.getElementById('kpi-vs-sub').textContent='VJ: €'+fmt(totalVJ);
document.getElementById('kpi-ist').textContent='€'+fmt(totalIst);
document.getElementById('kpi-ist-sub').textContent=istRows.length+' Tage verfügbar';
document.getElementById('s-growth').textContent=fmtPct(META.growthRate-1);
document.getElementById('s-holiday').textContent=META.holidayMult.toFixed(3)+'x';
document.getElementById('s-sf').textContent=META.sfMult.toFixed(3)+'x';
document.getElementById('s-sat').textContent=parseFloat(META.dayMultipliers['5']).toFixed(2);
document.getElementById('s-mon').textContent=parseFloat(META.dayMultipliers['0']).toFixed(2);
document.getElementById('s-fdays').textContent=RAW.filter(r=>r.feiertag).length;
const dayBarsEl=document.getElementById('day-bars');
const dayLabelsEl=document.getElementById('day-labels');
const mults=[0,1,2,3,4,5,6].map(i=>parseFloat(META.dayMultipliers[String(i)]));
const maxM=Math.max(...mults);
mults.forEach((m,i)=>{
  const bar=document.createElement('div');
  const h=Math.round((m/maxM)*46);
  bar.style.cssText=`flex:1;height:${h}px;background:${i===5?'var(--accent)':'var(--surface2)'};border-radius:2px 2px 0 0;border:1px solid var(--border);`;
  dayBarsEl.appendChild(bar);
  const lbl=document.createElement('div');
  lbl.style.cssText='flex:1;text-align:center;font-size:9px;color:var(--muted);';
  lbl.textContent=DAYS[i];
  dayLabelsEl.appendChild(lbl);
});
const months=[...new Set(RAW.map(r=>r.date.slice(0,7)))];
const mnames=['Jan','Feb','Mär','Apr','Mai','Jun','Jul','Aug','Sep','Okt','Nov','Dez'];
const pillsEl=document.getElementById('month-pills');
const allPill=document.createElement('button');
allPill.className='mpill active';allPill.textContent='Alle';
allPill.onclick=()=>setMonth('all',allPill);pillsEl.appendChild(allPill);
months.forEach(m=>{
  const idx=parseInt(m.slice(5))-1;
  const p=document.createElement('button');
  p.className='mpill';p.textContent=mnames[idx];
  p.onclick=()=>setMonth(m,p);pillsEl.appendChild(p);
});
function setMonth(m,el){activeMonth=m;document.querySelectorAll('.mpill').forEach(p=>p.classList.remove('active'));el.classList.add('active');renderTable();}
function setFilter(f,el){activeFilter=f;document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));el.classList.add('active');renderTable();}
function renderTable(){
  let rows=RAW;
  if(activeMonth!=='all')rows=rows.filter(r=>r.date.startsWith(activeMonth));
  if(activeFilter==='feiertag')rows=rows.filter(r=>r.feiertag);
  else if(activeFilter==='schulferien')rows=rows.filter(r=>r.schulferien);
  else if(activeFilter==='weekend')rows=rows.filter(r=>r.dow===5||r.dow===6);
  else if(activeFilter==='ist')rows=rows.filter(r=>r.ist);
  const tbody=document.getElementById('day-tbody');
  tbody.innerHTML=rows.map(r=>{
    const d=new Date(r.date);
    const dateStr=d.toLocaleDateString('de-DE',{day:'2-digit',month:'2-digit',year:'numeric'});
    const fcVsVJ=r.vj?(r.forecast-r.vj)/r.vj:null;
    const istVsFC=r.ist?(r.ist-r.forecast)/r.forecast:null;
    let typeTag=`<span class="tag-d">Werktag</span>`;
    if(r.feiertag)typeTag=`<span class="tag-f">${r.feiertag}</span>`;
    else if(r.schulferien)typeTag=`<span class="tag-s">${r.schulferien}</span>`;
    else if(r.dow===5||r.dow===6)typeTag=`<span class="tag-w">Wochenende</span>`;
    const istCell=r.ist?`<span class="dot-ist"></span>${fmt(r.ist)}`:'<span style="color:var(--muted)">—</span>';
    return `<tr><td>${dateStr}</td><td style="color:var(--muted)">${r.dowName}</td><td>${typeTag}</td>
      <td class="num">${fmt(r.vj)}</td>
      <td class="num" style="color:var(--accent);font-weight:500">${fmt(r.forecast)}</td>
      <td class="num">${istCell}</td>
      <td class="num ${fcVsVJ!=null?(fcVsVJ>=0?'up':'down'):''}">${fcVsVJ!=null?fmtPct(fcVsVJ):'—'}</td>
      <td class="num ${istVsFC!=null?(istVsFC>=0?'up':'down'):''}">${istVsFC!=null?fmtPct(istVsFC):'—'}</td>
    </tr>`;
  }).join('');
}
renderTable();
const yearCtx=document.getElementById('yearChart').getContext('2d');
const weeks={};
RAW.forEach(r=>{
  const m=r.date.slice(0,7);
  if(!weeks[m])weeks[m]={vj:0,fc:0,ist:0,hasIst:false,days:0};
  weeks[m].vj+=r.vj||0;weeks[m].fc+=r.forecast;
  if(r.ist){weeks[m].ist+=r.ist;weeks[m].hasIst=true;}
  weeks[m].days++;
});
const wkKeys=Object.keys(weeks).sort();
new Chart(yearCtx,{type:'line',data:{labels:wkKeys.map(k=>{const[,m]=k.split('-');return mnames[parseInt(m)-1];}),datasets:[
  {label:'Vorjahr',data:wkKeys.map(k=>Math.round(weeks[k].vj/weeks[k].days)),borderColor:'#3d4455',borderWidth:1.5,pointRadius:0,tension:0.4,fill:false,borderDash:[3,3]},
  {label:'Forecast',data:wkKeys.map(k=>Math.round(weeks[k].fc/weeks[k].days)),borderColor:'#e8f04a',borderWidth:2,pointRadius:3,pointBackgroundColor:'#e8f04a',tension:0.4,fill:false},
  {label:'Ist',data:wkKeys.map(k=>weeks[k].hasIst?Math.round(weeks[k].ist/weeks[k].days):null),borderColor:'#4af0a8',borderWidth:2,pointRadius:4,pointBackgroundColor:'#4af0a8',tension:0.4,fill:false,spanGaps:false},
]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: €${ctx.parsed.y.toLocaleString('de-DE')}`}}},scales:{x:{ticks:{color:'#6b7280',font:{size:10,family:'DM Mono'}},grid:{color:'rgba(37,42,56,.5)'}},y:{ticks:{color:'#6b7280',font:{size:10,family:'DM Mono'},callback:v=>'€'+Math.round(v).toLocaleString('de-DE')},grid:{color:'rgba(37,42,56,.5)'}}}}});
const mCtx=document.getElementById('monthChart').getContext('2d');
new Chart(mCtx,{type:'bar',data:{labels:MONTHLY.map(m=>m.month),datasets:[
  {label:'Vorjahr',data:MONTHLY.map(m=>m.vj),backgroundColor:'rgba(61,68,85,.7)',borderRadius:3},
  {label:'Forecast',data:MONTHLY.map(m=>m.forecast),backgroundColor:'rgba(232,240,74,.8)',borderRadius:3},
  {label:'Ist',data:MONTHLY.map(m=>m.ist),backgroundColor:'rgba(74,240,168,.85)',borderRadius:3},
]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false},tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: €${ctx.parsed.y.toLocaleString('de-DE')}`}}},scales:{x:{ticks:{color:'#6b7280',font:{size:10,family:'DM Mono'}},grid:{display:false}},y:{ticks:{color:'#6b7280',font:{size:10,family:'DM Mono'},callback:v=>'€'+(v/1000).toFixed(0)+'k'},grid:{color:'rgba(37,42,56,.5)'}}}}});
function exportCSV(){
  const hdr='Datum,Tag,Feiertag,Schulferien,Vorjahr €,Forecast €,Ist €,FC vs VJ %,Ist vs FC %';
  const rows=RAW.map(r=>{
    const fcVsVJ=r.vj?(((r.forecast-r.vj)/r.vj)*100).toFixed(1):'';
    const istVsFC=r.ist?(((r.ist-r.forecast)/r.forecast)*100).toFixed(1):'';
    return[r.date,r.dowName,r.feiertag||'',r.schulferien||'',r.vj||'',r.forecast,r.ist||'',fcVsVJ,istVsFC].join(',');
  });
  const blob=new Blob([[hdr,...rows].join('\n')],{type:'text/csv'});
  const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download='forecast___STORE_ID_____YEAR__.csv';a.click();
}
</script>
</body>
</html>"""


def build_html(results, monthly, meta, store_id, year):
    html = HTML_TEMPLATE
    html = html.replace("__STORE_ID__", str(store_id))
    html = html.replace("__YEAR__",     str(year))
    html = html.replace("__RESULTS__",  json.dumps(results,  separators=(",", ":")))
    html = html.replace("__MONTHLY__",  json.dumps(monthly,  separators=(",", ":")))
    html = html.replace("__META__",     json.dumps(meta,     separators=(",", ":")))
    return html


#  Main 

def main():
    # Resolve file path
    if len(sys.argv) > 1:
        xlsx_path = Path(sys.argv[1])
    else:
        xlsx_path = Path(DEFAULT_FILE)

    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        print(f"Usage: python {Path(__file__).name} path/to/your_file.xlsx")
        sys.exit(1)

    print(f"Loading: {xlsx_path}")

    # Load all data (open workbook separately each time — read_only streams can't rewind)
    print("  → TagesUmsatzVorjahr ...")
    vorjahr   = get_vorjahr(load_workbook(xlsx_path))
    print(f"     {len(vorjahr)} days loaded")

    print(f"  → Feiertage (state: {STATE_CODE}) ...")
    holidays  = get_holidays(load_workbook(xlsx_path), STATE_CODE)
    print(f"     {len(holidays)} holidays: {list(holidays.values())}")

    print(f"  → Schulferien (state: {STATE_CODE}) ...")
    schulferien = get_schulferien(load_workbook(xlsx_path), STATE_CODE)
    print(f"     {len(schulferien)} school-break days")

    print("  → Umsatz Forecast (Ist values) ...")
    ist_data  = get_ist_values(load_workbook(xlsx_path))
    ist_count = sum(1 for v in ist_data.values() if v.get("ist"))
    print(f"     {ist_count} days with actual (Ist) values")

    # Build model
    print("\nBuilding forecast model ...")
    day_mult, day_avgs, overall_avg, holiday_mult, sf_mult = build_model(
        vorjahr, holidays, schulferien
    )

    growth_rate = GROWTH_RATE or compute_growth_rate(ist_data, vorjahr)
    print(f"  Growth rate  : {(growth_rate-1)*100:+.1f}%")
    print(f"  Holiday mult : {holiday_mult:.3f}x")
    print(f"  Schulferien  : {sf_mult:.3f}x")
    print(f"  Day mults    : " +
          " | ".join(f"{'MoDiMiDoFrSaSo'.split()[i] if False else ['Mo','Di','Mi','Do','Fr','Sa','So'][i]}={v:.2f}"
                     for i, v in day_mult.items()))

    # Detect forecast year from vorjahr data (vorjahr year + 1)
    years = sorted({int(ds[:4]) for ds in vorjahr})
    forecast_year = max(years) + 1
    print(f"\nGenerating forecast for {forecast_year} ...")

    results = generate_forecast(
        forecast_year, vorjahr, holidays, schulferien, ist_data,
        day_mult, day_avgs, overall_avg, holiday_mult, sf_mult, growth_rate
    )
    monthly = monthly_summary(results)

    total_fc = sum(r["forecast"] for r in results)
    total_vj = sum(r["vj"] or 0 for r in results)
    print(f"  Total forecast : €{total_fc:,.0f}")
    print(f"  Total VJ       : €{total_vj:,.0f}")
    print(f"  vs VJ          : {(total_fc-total_vj)/total_vj*100:+.1f}%")

    meta = {
        "growthRate":      growth_rate,
        "holidayMult":     round(holiday_mult, 4),
        "sfMult":          round(sf_mult, 4),
        "dayMultipliers":  {str(k): round(v, 4) for k, v in day_mult.items()},
    }

    # Build HTML dashboard
    html = build_html(results, monthly, meta, STORE_ID, forecast_year)

    # Save next to the script
    out_html = xlsx_path.parent / f"forecast_{STORE_ID}_{forecast_year}.html"
    out_html.write_text(html, encoding="utf-8")
    print(f"\nDashboard saved → {out_html}")

    # Also export CSV
    out_csv = xlsx_path.parent / f"forecast_{STORE_ID}_{forecast_year}.csv"
    with open(out_csv, "w", encoding="utf-8") as f:
        f.write("Datum,Tag,Feiertag,Schulferien,Vorjahr_EUR,Forecast_EUR,Ist_EUR,FC_vs_VJ_pct,Ist_vs_FC_pct\n")
        for r in results:
            fc_vs_vj = f"{(r['forecast']-r['vj'])/r['vj']*100:.1f}" if r["vj"] else ""
            ist_vs_fc = f"{(r['ist']-r['forecast'])/r['forecast']*100:.1f}" if r["ist"] else ""
            f.write(",".join([
                r["date"], r["dowName"],
                r["feiertag"] or "", r["schulferien"] or "",
                str(r["vj"] or ""), str(r["forecast"]),
                str(r["ist"] or ""), fc_vs_vj, ist_vs_fc
            ]) + "\n")
    print(f"CSV saved      → {out_csv}")

    # Open in browser
    print("\nOpening dashboard in browser ...")
    webbrowser.open(out_html.resolve().as_uri())


if __name__ == "__main__":
    main()
